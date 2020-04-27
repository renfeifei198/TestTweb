from openpyxl import load_workbook, Workbook
from Config.VarConfig import *
from datetime import datetime
import os
from collections import OrderedDict
from Utils.Logger import log


class ExcelReader(object):
    @staticmethod
    def get_file_list():
        """
        取到script文件夹下所有excel的list
        """
        file_name_list = []
        try:
            file_list = os.listdir(EXCELPATH)
        except Exception as e:
            log.logger('ERROR', '获取excel文件失败：{}'.format(e))
            raise e
        for file in file_list:
            if os.path.splitext(file)[-1] == '.xlsx':
                file_name_list.append(os.path.splitext(file)[0])
        return file_name_list

    @staticmethod
    def get_sheet_value(filename, sheetname):
        global sh, wb
        datapath = os.path.join(EXCELPATH, filename) + '.xlsx'
        try:
            wb = load_workbook(datapath)
        except Exception as e:
            log.logger('WARNING', '未找到excel文件:{}'.format(e))
        try:
            sh = wb[sheetname]
        except Exception as e:
            log.logger('WARNING', '未找到sheetname:{}'.format(e))
        tablevalue = []
        rows = sh.max_row
        if rows < 2:
            return None
        else:
            for row in range(2, rows + 1):
                stepdict = []
                for i in range(1, 6):
                    stepdict.append(sh.cell(row, i).value)
                tablevalue.append(stepdict)
        return tablevalue


def get_sheet_names(filename):
    global sheetlist
    f = os.path.join(EXCELPATH, filename) + '.xlsx'
    try:
        wb = load_workbook(f)
        sheetlist = wb.sheetnames
    except Exception as e:
        log.logger('ERROR', '打开excel失败：{}'.format(e))
    return sheetlist


def save_sheet(filename, sheetname, testdata):
    datapath = os.path.join(EXCELPATH, filename) + '.xlsx'
    try:
        wb= load_workbook(datapath)
        log.logger('INFO', '存储到:{}'.format(filename))
    except Exception as e:
        log.logger('ERROR', '未找到指定excel')
        wb= Workbook()
        wb.save(datapath)
    try:
        sh= wb[sheetname]
        log.logger('INFO', '存储到:{}'.format(sheetname))
    except Exception as e:
        log.logger('INFO', '创建sheet:{}'.format(sheetname))
        sh = wb.create_sheet(sheetname)
    sh['A1'] = '操作步骤'
    sh['B1'] = '关键字'
    sh['C1'] = '定位方式'
    sh['D1'] = '定位对象'
    sh['E1'] = '输入的值'
    for i in range(0, len(testdata)):
        for j in range(0, 5):
            sh.cell(i+2, j+1).value = testdata[i][j]
    try:
        wb.save(datapath)
        log.logger('INFO', '保存成功!')
    except Exception as e:
        log.logger('ERROR', '保存失败，是不是excel打开着，没权限存储了：e'.format(e))
